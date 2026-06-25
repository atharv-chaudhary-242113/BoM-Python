"""
Handles all three output concerns:
  1. export_bom()        — writes final_bom.xlsx with grouped categories.
  2. export_exceptions() — writes exceptions.xlsx.
  3. safety_check()      — verifies output.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill

import config

# ── Dynamic Column Setup ───────────────────────────────────────────────────────
OUT_COL_SR = config.OUT_COL_SR

# Dynamically shift columns to accommodate CATEGORY
_shift = 0
if not hasattr(config, 'OUT_COL_CATEGORY'):
    OUT_COL_CATEGORY = config.OUT_COL_DESC
    _shift = 1
else:
    OUT_COL_CATEGORY = config.OUT_COL_CATEGORY

OUT_COL_DESC        = config.OUT_COL_DESC + _shift
OUT_COL_SPEC        = config.OUT_COL_SPEC + _shift
OUT_COL_MAKE        = config.OUT_COL_MAKE + _shift
OUT_COL_CATNO       = config.OUT_COL_CATNO + _shift
OUT_COL_UNIT        = config.OUT_COL_UNIT + _shift
OUT_PANEL_START_COL = config.OUT_PANEL_START_COL + _shift

# Overridden layout constants for unified single-row header
OUT_ROW_HEADER = 1
OUT_DATA_START = 2

# Core Constants
OUTPUT_DIR     = config.OUTPUT_DIR
OUTPUT_BOM     = config.OUTPUT_BOM
EXCEPTIONS_FILE= config.EXCEPTIONS_FILE
FILL_PINK      = config.FILL_PINK
FILL_GREEN     = config.FILL_GREEN
FONT_BLACK     = config.FONT_BLACK


# ── Shared style objects ───────────────────────────────────────────────────────

def _pink_fill() -> PatternFill:
    return PatternFill(start_color=FILL_PINK, end_color=FILL_PINK, fill_type="solid")

def _green_fill() -> PatternFill:
    return PatternFill(start_color=FILL_GREEN, end_color=FILL_GREEN, fill_type="solid")

def _black_font(bold: bool = True) -> Font:
    return Font(color=FONT_BLACK, bold=bold)

def _to_number(value):
    if value is None or value == "":
        return None
    try:
        f = float(value)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return value


# ── 1. Export final BoM ───────────────────────────────────────────────────────

def export_bom(
    all_panels: list[str],
    sorted_rows: list[dict],
    logger,
) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidated BoM"

    pink  = _pink_fill()
    green = _green_fill()
    bold_black = _black_font(bold=True)

    # ── Row 1: Static headers ─────────────────────────────────────────────────
    static_headers = {
        OUT_COL_SR:       "SNo",
        OUT_COL_CATEGORY: "CATEGORY",
        OUT_COL_DESC:     "DESCRIPTION",
        OUT_COL_SPEC:     "SPEC",
        OUT_COL_MAKE:     "MAKE",
        OUT_COL_CATNO:    "CAT NO.",
        OUT_COL_UNIT:     "UNIT",
    }

    for col, label in static_headers.items():
        cell = ws.cell(row=OUT_ROW_HEADER, column=col)
        cell.value = label
        cell.fill  = pink
        cell.font  = bold_black

    # ── Row 1: Panel names ────────────────────────────────────────────────────
    for i, panel_name in enumerate(all_panels):
        col  = OUT_PANEL_START_COL + i
        cell = ws.cell(row=OUT_ROW_HEADER, column=col)
        cell.value = panel_name
        cell.fill  = green
        cell.font  = bold_black

    # ── Data Rows ─────────────────────────────────────────────────────────────
    item_count = 1
    current_out_row = OUT_DATA_START
    last_category = None

    for row in sorted_rows:
        cat = row.get("CATEGORY", "Uncategorized")

        # Inject blank spacer row when category changes (except before the first category)
        if last_category is not None and cat != last_category:
            current_out_row += 1

        last_category = cat

        ws.cell(row=current_out_row, column=OUT_COL_SR).value = item_count
        ws.cell(row=current_out_row, column=OUT_COL_CATEGORY).value = cat
        ws.cell(row=current_out_row, column=OUT_COL_DESC).value = row["DESCRIPTION"]
        ws.cell(row=current_out_row, column=OUT_COL_SPEC).value = row["SPEC"]
        ws.cell(row=current_out_row, column=OUT_COL_MAKE).value = row["MAKE"]
        ws.cell(row=current_out_row, column=OUT_COL_CATNO).value = row["CAT NO."]
        ws.cell(row=current_out_row, column=OUT_COL_UNIT).value = row["UNIT"]

        pq = row.get("panel_quantities", {})
        for i, panel_name in enumerate(all_panels):
            col = OUT_PANEL_START_COL + i
            ws.cell(row=current_out_row, column=col).value = _to_number(pq.get(panel_name))

        item_count += 1
        current_out_row += 1

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, OUTPUT_BOM)
    wb.save(out_path)
    logger.info(f"Final BoM written → {out_path}  ({len(sorted_rows)} items, {len(all_panels)} panels)")
    return out_path


# ── 2. Export exceptions ──────────────────────────────────────────────────────

def export_exceptions(exceptions: list[dict], all_panels: list[str], logger) -> None:
    if not exceptions:
        logger.info("No exceptions — exceptions file not written.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exceptions"

    static_fields = ["DESCRIPTION", "SPEC", "MAKE", "UNIT", "CAT NO."]
    header = ["Sheet", "Row", "Issue", "Description", ""] + static_fields + all_panels
    ws.append(header)

    for exc in exceptions:
        raw = exc.get("raw_row")
        if raw:
            pq = raw.get("panel_quantities", {})
            raw_static = [
                str(raw.get("DESCRIPTION") or ""),
                str(raw.get("SPEC")        or ""),
                str(raw.get("MAKE")        or ""),
                str(raw.get("UNIT")        or ""),
                str(raw.get("CAT NO.")     or ""),
            ]
            raw_panels = [pq.get(p) for p in all_panels]
        else:
            raw_static = [""] * len(static_fields)
            raw_panels = [""] * len(all_panels)

        row = (
            [exc.get("sheet", ""), exc.get("row", ""),
             exc.get("issue", ""), exc.get("description", ""), ""]
            + raw_static
            + raw_panels
        )
        ws.append(row)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, EXCEPTIONS_FILE)
    wb.save(out_path)
    logger.info(f"Exceptions file written → {out_path}  ({len(exceptions)} exception(s))")


# ── 3. Safety check ───────────────────────────────────────────────────────────

def safety_check(
    all_panels: list[str],
    sorted_rows: list[dict],
    sheets_data: list[dict],
    logger,
) -> bool:
    logger.info("─── Safety check starting ───────────────────────────────────")
    passed = True

    source_panels: set[str] = set()
    for sheet in sheets_data:
        for _col, panel_name in sheet["panels"]:
            source_panels.add(panel_name)

    output_panels = set(all_panels)
    missing = source_panels - output_panels
    extra   = output_panels - source_panels

    if missing:
        logger.error(f"FAIL panel names: missing from output → {sorted(missing)}")
        passed = False
    else:
        logger.info(f"PASS panel names: all {len(source_panels)} panel(s) present in output.")

    if extra:
        logger.warning(f"NOTE: output contains panel(s) not seen in source → {sorted(extra)}")

    def _sum_qty(rows: list[dict], panel: str) -> float:
        total = 0.0
        for row in rows:
            v = row.get("panel_quantities", {}).get(panel)
            try:
                total += float(v) if v is not None and v != "" else 0.0
            except (TypeError, ValueError):
                pass
        return total

    all_source_rows: list[dict] = [r for s in sheets_data for r in s["data_rows"]]

    qty_ok = True
    for panel_name in source_panels:
        src_total = _sum_qty(all_source_rows, panel_name)
        out_total = _sum_qty(sorted_rows, panel_name)
        if abs(src_total - out_total) > 1e-9:
            logger.error(
                f"FAIL quantity mismatch [{panel_name}]: source={src_total}, output={out_total}"
            )
            qty_ok = False
            passed = False

    if qty_ok:
        logger.info("PASS quantities: all panel totals match source.")

        def _get_key(row: dict) -> str:
            cat_no = str(row.get("CAT NO.") or "").strip()
            if cat_no:
                return f"cat::{cat_no}"
            return f"desc::{str(row.get('DESCRIPTION', '')).strip()}"

        source_keys: set[str] = set()
        for row in all_source_rows:
            source_keys.add(_get_key(row))

        output_count = len(sorted_rows)

        if len(source_keys) == output_count:
            logger.info(f"PASS item count: {output_count} unique item(s) in output matches source.")
        else:
            logger.error(f"FAIL item count: source has {len(source_keys)} unique item(s), output has {output_count}.")
            passed = False

    return passed