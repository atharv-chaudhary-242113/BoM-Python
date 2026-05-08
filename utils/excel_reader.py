"""
Loads a .xlsx workbook and extracts BoM rows from every sheet that
contains a valid panel row (row 1, from col G) and a recognisable header
row (row 4). Category state resets per sheet.
"""

import openpyxl

from config import HEADER_ROW, STATIC_COLS
from utils.panel_detector import detect_panels
from utils.row_cleaner import clean_row

SRNO_COL = 1


def _build_header_map(ws) -> dict[str, int]:
    header_map: dict[str, int] = {}
    col = 1
    while True:
        cell = ws.cell(row=HEADER_ROW, column=col)
        if col > 150:
            break
        raw = cell.value
        if raw is not None:
            normalised = str(raw).strip().upper()
            if normalised in STATIC_COLS:
                header_map[normalised] = col
        col += 1
    return header_map


def _verify_srno_header(ws) -> bool:
    val = ws.cell(row=HEADER_ROW, column=SRNO_COL).value
    if val is None:
        return False
    return str(val).strip().upper() == "SNO"


def read_workbook(
    file_path: str,
    logger,
) -> tuple[list[dict], list[dict]]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets_data: list[dict] = []
    exceptions:  list[dict] = []

    def _is_empty(val):
        if val is None:
            return True
        text = str(val).strip().upper()
        return text in ("", "0", "0.0", "NONE", "NULL", "-")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # FIX: Reset per sheet so categories don't bleed from the previous sheet
        current_category = "Uncategorized"

        # ── 1. Detect panels ──────────────────────────────────────────────
        panels = detect_panels(ws)
        if not panels:
            logger.warning(f"Sheet '{sheet_name}': no panel names in row 1 from col G. Skipping.")
            exceptions.append({
                "sheet":       sheet_name,
                "row":         1,
                "issue":       "No panels detected",
                "description": "Row 1 from column G is empty. Sheet skipped.",
                "raw_row":     None,
            })
            continue

        # ── 2. Verify SNo is in column A ──────────────────────────────────
        if not _verify_srno_header(ws):
            actual = ws.cell(row=HEADER_ROW, column=SRNO_COL).value
            logger.warning(
                f"Sheet '{sheet_name}': expected 'SNo' in A4, "
                f"found {repr(actual)!r}. Skipping."
            )
            exceptions.append({
                "sheet":       sheet_name,
                "row":         4,
                "issue":       "SNo header missing from column A",
                "description": f"A4 contains {repr(actual)} instead of 'SNo'. Sheet skipped.",
                "raw_row":     None,
            })
            continue

        # ── 3. Locate static column positions ─────────────────────────────
        header_map = _build_header_map(ws)

        if "DESCRIPTION" not in header_map:
            logger.warning(f"Sheet '{sheet_name}': DESCRIPTION header not found in row 4. Skipping.")
            exceptions.append({
                "sheet":       sheet_name,
                "row":         4,
                "issue":       "DESCRIPTION header missing",
                "description": "Row 4 has no DESCRIPTION header. Sheet skipped.",
                "raw_row":     None,
            })
            continue

        # ── 4. Extract data rows ───────────────────────────────────────────
        data_rows: list[dict] = []
        panel_col_indices = {panel_name: col_idx for col_idx, panel_name in panels}

        row_num = HEADER_ROW + 1
        empty_streak = 0

        while True:
            raw: dict = {}
            for col_name in STATIC_COLS:
                col_idx = header_map.get(col_name)
                raw[col_name] = ws.cell(row=row_num, column=col_idx).value if col_idx else None

            raw["panel_quantities"] = {}
            for panel_name, col_idx in panel_col_indices.items():
                raw["panel_quantities"][panel_name] = ws.cell(row=row_num, column=col_idx).value

            srno_val = ws.cell(row=row_num, column=SRNO_COL).value

            srno_empty = _is_empty(srno_val)
            desc_empty = _is_empty(raw.get("DESCRIPTION"))
            cat_empty  = _is_empty(raw.get("CAT NO."))
            spec_empty = _is_empty(raw.get("SPEC"))
            make_empty = _is_empty(raw.get("MAKE"))
            unit_empty = _is_empty(raw.get("UNIT"))

            has_qty = any(not _is_empty(v) for v in raw["panel_quantities"].values())

            # ── Junk / Footer Interception ────────────────────────────────
            # Only trigger when SNo is also absent — prevents swallowing
            # numbered category-header rows (e.g. "39 | WIRE | ...")
            if srno_empty and desc_empty and cat_empty:
                if not has_qty:
                    empty_streak += 1
                    if empty_streak >= 5:
                        break
                else:
                    empty_streak = 0
                row_num += 1
                continue

            empty_streak = 0

            # ── Category Interception ─────────────────────────────────────
            # FIX: srno_empty removed. In this workbook, all category headers
            # after the first one carry a sequential SNo (e.g. "39 | WIRE").
            # The reliable signals are: has a description, but absolutely no
            # CAT NO., SPEC, MAKE, UNIT, and no panel quantities.
            is_category = (
                not desc_empty
                and cat_empty
                and spec_empty
                and make_empty
                and unit_empty
                and not has_qty
            )

            if is_category:
                current_category = str(raw.get("DESCRIPTION")).strip()
                logger.debug(
                    f"Sheet '{sheet_name}' row {row_num}: "
                    f"category → '{current_category}'"
                )
                row_num += 1
                continue

            # ── Standard Item Extraction ───────────────────────────────────
            raw["CATEGORY"] = current_category

            ok, cleaned, exc = clean_row(raw, sheet_name, row_num)
            if ok:
                data_rows.append(cleaned)
            elif exc:
                exceptions.append(exc)

            row_num += 1

        logger.info(
            f"Sheet '{sheet_name}': {len(panels)} panel(s), "
            f"{len(data_rows)} valid item(s) extracted."
        )

        sheets_data.append({
            "sheet_name": sheet_name,
            "panels":     panels,
            "data_rows":  data_rows,
        })

    wb.close()
    return sheets_data, exceptions